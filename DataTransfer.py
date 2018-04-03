import openpyxl as xl
# In Progress Currently
# Idea: Add parsing through all folder location and transferring automatically

print("File Location must be double \ not single Include \FileName.ext")
path1 = input("Location: Original\n")
path2 = input("Location: Copy to\n")
wb1 = xl.load_workbook(filename=path1)
wb2 = xl.load_workbook(filename=path2)

menuSelect = input("Which Option? 1. Copy columns. 2. Copy entire\n")
#TESTING FUNCTIONS#
if menuSelect == '1':
    ws1 = wb1.active
    ws2 = wb2.active
    columnIndex = input("Which column? 0-indexing\n")
    actualIndex = input("What column? (1-index\n")

    col_numb = ws1.columns[columnIndex]
    for idx, cell in enumerate(col_numb, 1):
        ws.cell(row=idx, column=actualIndex).value = cell.value
    wb2.save(path2)
    #error - fix this to copy and also parse through column.
    #https://stackoverflow.com/questions/32173053/copy-columns-from-workbook-paste-in-second-sheet-of-second-workbook-openpyxl
    #https://openpyxl.readthedocs.io/en/stable/tutorial.html#saving-to-a-file
elif menuSelect == '2':
    ws1 = wb1.worksheets[0]
    ws2 = wb2.create_sheet(ws1.title)

    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    wb2.save(path2)




#wb1 - a = wb2 - a
def CopyColumn():
    ws1 = wb1.active
    ws2 = wb2.active
    columnIndex = input("Which column? 0-indexing\n")
    actualIndex = input("What column? (1-index\n")

    ws1.columns[columnName]
    for idx, cell in enumerate(columnIndex, 1):
        ws.cell(row=idx, column=actualIndex).value = cell.value
    wb2.save(path2)

#wb1 = wb2
def  CopyAll():
    ws1 = wb1.worksheets[0]
    ws2 = wb2.create_sheet(ws1.title)

    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    wb2.save(path2)

#C:\\Users\\logan\\Documents\\Python\\DataBook1.xlsx
